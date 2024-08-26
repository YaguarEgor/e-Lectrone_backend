from fastapi import APIRouter
from openpyxl import load_workbook




router = APIRouter(
    prefix="/tutors",
    tags=["Tutor"]
)

def get_all_tutors():
    wb = load_workbook("./tutors.xlsx")
    sheet = wb.get_sheet_by_name("Tutors")


    d = []
    for i in range(2, sheet.max_row + 1):
        prices_s = sheet[f'F{i}'].value.strip().split(";")
        prices = dict()
        for j in prices_s:
            prices[j[:j.index(":")]] = j[j.index(":")+1:]
        d.append({
            "id": sheet[f'A{i}'].value,
            "name": sheet[f'B{i}'].value,
            "description": sheet[f'C{i}'].value,
            "avatar": sheet[f'D{i}'].value,
            "subjects": list(sheet[f'E{i}'].value.split()),
            "prices": prices,
            "contact": sheet[f'G{i}'].value
        })
    return d


@router.get('/')
async def get_tutors():
    return get_all_tutors()

@router.get('/{id}')
async def get_tutor_info(id):
    tutors_l = get_all_tutors()
    for i in tutors_l:
        if int(i['id']) == int(id):
            return i
    return None

