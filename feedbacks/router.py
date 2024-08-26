from fastapi import APIRouter
from openpyxl import load_workbook




router = APIRouter(
    prefix="/feedbacks",
    tags=["Feedback"]
)

@router.get('/')
async def get_feedbacks():
    wb = load_workbook("./feedbacks.xlsx")
    sheet = wb.get_sheet_by_name("Feedbacks")
    d = []
    for i in range(1, sheet.max_row+1):
        d.append({
            "id": sheet[f'A{i}'].value,
            "title": sheet[f'B{i}'].value,
            "all_text": sheet[f'C{i}'].value,
            "author": sheet[f'D{i}'].value
        })
    return d

